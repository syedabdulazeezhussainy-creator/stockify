import os
import sqlite3
import bcrypt
import csv
from datetime import datetime, timedelta
from functools import wraps
import glob
from flask import Flask, render_template, request, redirect, session, flash, send_file, url_for
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO, StringIO
import qrcode
import base64

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "supersecretkey123")
app.permanent_session_lifetime = 3600

# ------------------ HELPER FUNCTIONS ------------------
def get_db():
    db_name = session.get("db_name")
    if not db_name:
        return sqlite3.connect("instance/default.db")
    return sqlite3.connect(f"instance/{db_name}")

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            flash("Please log in first.", "warning")
            return redirect("/staff-login")
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Access denied. Admin only.", "danger")
            return redirect("/dashboard")
        return f(*args, **kwargs)
    return decorated

def log_activity(action, details=""):
    if "user" not in session:
        return
    with get_db() as conn:
        conn.execute("""
            INSERT INTO activity_log (user_id, username, action, details, timestamp)
            VALUES (?, ?, ?, ?, ?)
        """, (
            session.get("user_id"),
            session.get("user"),
            action,
            details,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

def get_company_list():
    db_files = glob.glob("instance/stockify_*.db")
    companies = []
    for f in db_files:
        name = f[18:-3].replace('_', ' ').title()
        companies.append((os.path.basename(f), name))
    return companies

def init_company_db(db_name):
    with sqlite3.connect(f"instance/{db_name}") as conn:
        # Users
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                password BLOB,
                role TEXT DEFAULT 'staff',
                email TEXT
            )
        """)
        try:
            conn.execute("ALTER TABLE users ADD COLUMN email TEXT")
        except:
            pass

        # Company info
        conn.execute("""
            CREATE TABLE IF NOT EXISTS company (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                gst TEXT,
                category TEXT
            )
        """)

        # Categories
        conn.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE
            )
        """)

        # Suppliers
        conn.execute("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                phone TEXT
            )
        """)

        # Products
        conn.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                sku TEXT UNIQUE,
                category_id INTEGER,
                supplier_id INTEGER,
                price REAL,
                stock INTEGER DEFAULT 0
            )
        """)

        # Stock movements
        conn.execute("""
            CREATE TABLE IF NOT EXISTS stock_in (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                quantity INTEGER,
                date TEXT,
                notes TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS stock_out (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                quantity INTEGER,
                date TEXT,
                notes TEXT
            )
        """)

        # Sales
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                customer_id INTEGER,
                quantity INTEGER,
                total REAL,
                date TEXT,
                warranty_months INTEGER DEFAULT 0,
                payment_method TEXT
            )
        """)
        try:
            conn.execute("ALTER TABLE sales ADD COLUMN payment_method TEXT")
        except:
            pass

        # Customers
        conn.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                phone TEXT UNIQUE,
                email TEXT,
                address TEXT,
                points INTEGER DEFAULT 0
            )
        """)
        try:
            conn.execute("ALTER TABLE customers ADD COLUMN points INTEGER DEFAULT 0")
        except:
            pass

        # Warranties
        conn.execute("""
            CREATE TABLE IF NOT EXISTS warranties (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER,
                product_id INTEGER,
                customer_id INTEGER,
                start_date TEXT,
                end_date TEXT,
                status TEXT DEFAULT 'active'
            )
        """)

        # Services
        conn.execute("""
            CREATE TABLE IF NOT EXISTS services (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                customer_id INTEGER,
                issue TEXT,
                service_date TEXT,
                resolution TEXT,
                cost REAL,
                status TEXT DEFAULT 'pending'
            )
        """)

        # Returns
        conn.execute("""
            CREATE TABLE IF NOT EXISTS returns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER,
                product_id INTEGER,
                customer_id INTEGER,
                reason TEXT,
                refund_amount REAL,
                return_date TEXT,
                status TEXT DEFAULT 'pending'
            )
        """)

        # Activity Log
        conn.execute("""
            CREATE TABLE IF NOT EXISTS activity_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                action TEXT,
                details TEXT,
                timestamp TEXT
            )
        """)

        # Default admin
        admin = conn.execute("SELECT * FROM users WHERE username='admin'").fetchone()
        if not admin:
            hashed = bcrypt.hashpw("admin".encode(), bcrypt.gensalt())
            conn.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                         ("admin", hashed, "admin"))

        # Default categories
        default_cats = ["Electronics", "Clothing", "Vehicles", "Hardware", "Food Items"]
        for cat in default_cats:
            conn.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (cat,))

# ------------------ ROUTES ------------------
@app.route("/")
def splash():
    return render_template("splash.html")

@app.route("/admin-login")
def admin_login():
    return render_template("admin_login.html")

@app.route("/staff-login")
def staff_login():
    companies = get_company_list()
    return render_template("staff_login.html", companies=companies)

@app.route("/login", methods=["POST"])
def login():
    if "company_name" in request.form:
        # Admin login
        company = request.form["company_name"].strip()
        gst = request.form["gst"].strip()
        category = request.form["category"].strip()
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        db_name = f"stockify_{company.replace(' ', '_').lower()}.db"
        session["db_name"] = db_name
        init_company_db(db_name)

        with sqlite3.connect(f"instance/{db_name}") as conn:
            conn.execute("DELETE FROM company")
            conn.execute("INSERT INTO company (name, gst, category) VALUES (?, ?, ?)",
                         (company, gst, category))
            user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()

        if user and bcrypt.checkpw(password.encode(), user[2]):
            session.permanent = True
            session["user"] = username
            session["user_id"] = user[0]
            session["company"] = company
            session["role"] = user[3]
            log_activity("LOGIN", "Admin login successful")
            flash(f"Welcome admin {username}!", "success")
            return redirect("/dashboard")
        else:
            flash("Invalid admin credentials", "danger")
            return redirect("/admin-login")
    else:
        # Staff login
        db_name = request.form["company_db"]
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        session["db_name"] = db_name

        with sqlite3.connect(f"instance/{db_name}") as conn:
            user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
            if user and bcrypt.checkpw(password.encode(), user[2]):
                company = conn.execute("SELECT name FROM company").fetchone()
                session.permanent = True
                session["user"] = username
                session["user_id"] = user[0]
                session["company"] = company[0] if company else "Unknown"
                session["role"] = user[3]
                log_activity("LOGIN", "Staff login successful")
                flash(f"Welcome {username}!", "success")
                return redirect("/dashboard")
            else:
                flash("Invalid staff credentials", "danger")
                return redirect("/staff-login")

@app.route("/dashboard")
@login_required
def dashboard():
    with get_db() as conn:
        total_products = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        total_sales = conn.execute("SELECT SUM(total) FROM sales").fetchone()[0] or 0
        low_stock = conn.execute("SELECT COUNT(*) FROM products WHERE stock < 5").fetchone()[0]
        pending_services = conn.execute("SELECT COUNT(*) FROM services WHERE status='pending'").fetchone()[0]
        active_warranties = conn.execute("SELECT COUNT(*) FROM warranties WHERE status='active'").fetchone()[0]
        today = datetime.now().strftime("%Y-%m-%d")
        today_sales = conn.execute("SELECT SUM(total) FROM sales WHERE date=?", (today,)).fetchone()[0] or 0
        expiring = conn.execute("""
            SELECT COUNT(*) FROM warranties
            WHERE status='active' AND date(end_date) BETWEEN date('now') AND date('now', '+7 days')
        """).fetchone()[0]
        # Reorder predictions (simple)
        predictions = conn.execute("""
            SELECT p.id, p.name, p.stock,
                   COALESCE(ROUND(AVG(s.quantity) / 30.0, 2), 0) AS daily_avg
            FROM products p
            LEFT JOIN sales s ON p.id = s.product_id AND s.date >= date('now', '-30 days')
            GROUP BY p.id
        """).fetchall()
        reorder_items = [row for row in predictions if row[3] > 0 and row[2] < (row[3] * 7)]
    return render_template("dashboard.html",
                           total_products=total_products,
                           total_sales=total_sales,
                           low_stock_count=low_stock,
                           pending_services=pending_services,
                           active_warranties=active_warranties,
                           today_sales=today_sales,
                           expiring_warranties=expiring,
                           reorder_items=reorder_items)

# ------------------ ADMIN ONLY ROUTES ------------------
@app.route("/products", methods=["GET", "POST"])
@login_required
@admin_required
def products():
    if request.method == "POST":
        name = request.form["name"]
        price = float(request.form["price"])
        category = request.form["category"]
        supplier = request.form["supplier"]
        sku = name[:3].upper() + str(int(datetime.now().timestamp()))[-4:]
        with get_db() as conn:
            conn.execute("""
                INSERT INTO products (name, sku, category_id, supplier_id, price, stock)
                VALUES (?, ?, ?, ?, ?, 0)
            """, (name, sku, category, supplier, price))
        log_activity("ADD_PRODUCT", f"Added product: {name}")
        flash("Product added successfully!", "success")

    with get_db() as conn:
        categories = conn.execute("SELECT * FROM categories").fetchall()
        suppliers = conn.execute("SELECT * FROM suppliers").fetchall()
        products = conn.execute("""
            SELECT p.id, p.name, p.sku, p.price, p.stock,
                   c.name AS cat_name, s.name AS sup_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN suppliers s ON p.supplier_id = s.id
        """).fetchall()
    return render_template("products.html", products=products,
                           categories=categories, suppliers=suppliers)

@app.route("/add-stock/<int:product_id>", methods=["POST"])
@login_required
@admin_required
def add_stock(product_id):
    qty = int(request.form["quantity"])
    notes = request.form.get("notes", "")
    date = datetime.now().strftime("%Y-%m-%d")
    with get_db() as conn:
        conn.execute("INSERT INTO stock_in (product_id, quantity, date, notes) VALUES (?, ?, ?, ?)",
                     (product_id, qty, date, notes))
        conn.execute("UPDATE products SET stock = stock + ? WHERE id = ?", (qty, product_id))
    log_activity("STOCK_IN", f"Added {qty} units to product ID {product_id}")
    flash(f"Added {qty} units to stock.", "success")
    return redirect("/products")

@app.route("/remove-stock/<int:product_id>", methods=["POST"])
@login_required
@admin_required
def remove_stock(product_id):
    qty = int(request.form["quantity"])
    notes = request.form.get("notes", "")
    date = datetime.now().strftime("%Y-%m-%d")
    with get_db() as conn:
        current = conn.execute("SELECT stock FROM products WHERE id=?", (product_id,)).fetchone()
        if current and current[0] >= qty:
            conn.execute("INSERT INTO stock_out (product_id, quantity, date, notes) VALUES (?, ?, ?, ?)",
                         (product_id, qty, date, notes))
            conn.execute("UPDATE products SET stock = stock - ? WHERE id = ?", (qty, product_id))
            log_activity("STOCK_OUT", f"Removed {qty} units from product ID {product_id}")
            flash(f"Removed {qty} units from stock.", "success")
        else:
            flash("Insufficient stock!", "danger")
    return redirect("/products")

@app.route("/suppliers", methods=["GET", "POST"])
@login_required
@admin_required
def suppliers():
    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        with get_db() as conn:
            conn.execute("INSERT INTO suppliers (name, phone) VALUES (?, ?)", (name, phone))
        log_activity("ADD_SUPPLIER", f"Added supplier: {name}")
        flash("Supplier added successfully!", "success")

    with get_db() as conn:
        suppliers = conn.execute("SELECT * FROM suppliers").fetchall()
    return render_template("suppliers.html", suppliers=suppliers)

@app.route("/admin/users", methods=["GET", "POST"])
@login_required
@admin_required
def admin_users():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        role = request.form["role"]
        email = request.form.get("email", "")
        hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
        try:
            with get_db() as conn:
                conn.execute("INSERT INTO users (username, password, role, email) VALUES (?, ?, ?, ?)",
                             (username, hashed, role, email))
            log_activity("ADD_USER", f"Created user: {username} with role {role}")
            flash(f"User {username} created.", "success")
        except sqlite3.IntegrityError:
            flash("Username already exists.", "danger")

    with get_db() as conn:
        users = conn.execute("SELECT id, username, role, email FROM users").fetchall()
    return render_template("admin_users.html", users=users)

@app.route("/activity-log")
@login_required
@admin_required
def activity_log():
    with get_db() as conn:
        logs = conn.execute("""
            SELECT username, action, details, timestamp
            FROM activity_log
            ORDER BY timestamp DESC LIMIT 100
        """).fetchall()
    return render_template("activity_log.html", logs=logs)

# ------------------ SHARED ROUTES ------------------
@app.route("/sales", methods=["GET", "POST"])
@login_required
def sales():
    if request.method == "POST":
        product_id = request.form["product"]
        qty = int(request.form["qty"])
        customer_id = request.form.get("customer")
        warranty_months = int(request.form.get("warranty_months", 0))
        payment_method = request.form.get("payment_method", "cash")

        with get_db() as conn:
            product = conn.execute("SELECT price, stock FROM products WHERE id=?", (product_id,)).fetchone()
            if not product:
                flash("Product not found.", "danger")
                return redirect("/sales")
            price, stock = product
            if qty > stock:
                qty = stock
                flash(f"Quantity reduced to available stock ({stock}).", "warning")
            total = qty * price
            date = datetime.now().strftime("%Y-%m-%d")
            cur = conn.execute("""
                INSERT INTO sales (product_id, customer_id, quantity, total, date, warranty_months, payment_method)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (product_id, customer_id, qty, total, date, warranty_months, payment_method))
            sale_id = cur.lastrowid
            conn.execute("UPDATE products SET stock = stock - ? WHERE id = ?", (qty, product_id))

            if warranty_months > 0 and customer_id:
                start_date = date
                end_date = (datetime.now() + timedelta(days=30 * warranty_months)).strftime("%Y-%m-%d")
                conn.execute("""
                    INSERT INTO warranties (sale_id, product_id, customer_id, start_date, end_date, status)
                    VALUES (?, ?, ?, ?, ?, 'active')
                """, (sale_id, product_id, customer_id, start_date, end_date))

            if customer_id:
                points_earned = int(total / 100)
                conn.execute("UPDATE customers SET points = points + ? WHERE id = ?", (points_earned, customer_id))
                flash(f"🎁 {points_earned} loyalty points added!", "success")

            if payment_method != "cash":
                flash(f"💳 Payment via {payment_method.upper()} simulated.", "info")

        log_activity("SALE", f"Sale of {qty} units of product ID {product_id} to customer {customer_id}")
        flash("Sale recorded successfully!", "success")

    with get_db() as conn:
        products = conn.execute("SELECT id, name, price, stock FROM products").fetchall()
        customers = conn.execute("SELECT id, name, phone, points FROM customers").fetchall()
        recent_sales = conn.execute("""
            SELECT s.id, s.date, p.name, c.name, s.quantity, s.total, s.payment_method
            FROM sales s
            JOIN products p ON s.product_id = p.id
            LEFT JOIN customers c ON s.customer_id = c.id
            ORDER BY s.date DESC LIMIT 10
        """).fetchall()
    return render_template("sales.html", products=products, customers=customers, recent_sales=recent_sales)

@app.route("/customers", methods=["GET", "POST"])
@login_required
def customers():
    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        email = request.form.get("email", "")
        address = request.form.get("address", "")
        try:
            with get_db() as conn:
                conn.execute("INSERT INTO customers (name, phone, email, address) VALUES (?, ?, ?, ?)",
                             (name, phone, email, address))
            log_activity("ADD_CUSTOMER", f"Added customer: {name}")
            flash("Customer added successfully!", "success")
        except sqlite3.IntegrityError:
            flash("Phone number already exists.", "danger")

    with get_db() as conn:
        customers = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    return render_template("customers.html", customers=customers)

@app.route("/warranty")
@login_required
def warranty():
    with get_db() as conn:
        warranties = conn.execute("""
            SELECT w.id, p.name AS product, c.name AS customer,
                   w.start_date, w.end_date, w.status
            FROM warranties w
            JOIN products p ON w.product_id = p.id
            JOIN customers c ON w.customer_id = c.id
            ORDER BY w.end_date
        """).fetchall()
    return render_template("warranty.html", warranties=warranties)

@app.route("/service", methods=["GET", "POST"])
@login_required
def service():
    if request.method == "POST":
        product_id = request.form["product"]
        customer_id = request.form["customer"]
        issue = request.form["issue"]
        service_date = datetime.now().strftime("%Y-%m-%d")
        with get_db() as conn:
            conn.execute("""
                INSERT INTO services (product_id, customer_id, issue, service_date, status)
                VALUES (?, ?, ?, ?, 'pending')
            """, (product_id, customer_id, issue, service_date))
        log_activity("SERVICE_CREATED", f"Service logged for product ID {product_id}")
        flash("Service request logged.", "success")

    with get_db() as conn:
        products = conn.execute("SELECT id, name FROM products").fetchall()
        customers = conn.execute("SELECT id, name FROM customers").fetchall()
        services = conn.execute("""
            SELECT s.id, p.name, c.name, s.issue, s.service_date, s.status
            FROM services s
            JOIN products p ON s.product_id = p.id
            JOIN customers c ON s.customer_id = c.id
            ORDER BY s.service_date DESC
        """).fetchall()
    return render_template("service.html", products=products, customers=customers, services=services)

@app.route("/service/update/<int:service_id>", methods=["POST"])
@login_required
def update_service(service_id):
    status = request.form["status"]
    resolution = request.form.get("resolution", "")
    cost = float(request.form.get("cost", 0))
    with get_db() as conn:
        conn.execute("""
            UPDATE services SET status=?, resolution=?, cost=?
            WHERE id=?
        """, (status, resolution, cost, service_id))
    log_activity("SERVICE_UPDATED", f"Service {service_id} updated to {status}")
    flash("Service updated.", "success")
    return redirect("/service")

@app.route("/returns", methods=["GET", "POST"])
@login_required
def returns():
    if request.method == "POST":
        sale_id = request.form["sale_id"]
        reason = request.form["reason"]
        refund_amount = float(request.form["refund_amount"])
        return_date = datetime.now().strftime("%Y-%m-%d")

        with get_db() as conn:
            sale = conn.execute("SELECT product_id, customer_id FROM sales WHERE id=?", (sale_id,)).fetchone()
            if sale:
                product_id, customer_id = sale
                conn.execute("""
                    INSERT INTO returns (sale_id, product_id, customer_id, reason, refund_amount, return_date, status)
                    VALUES (?, ?, ?, ?, ?, ?, 'completed')
                """, (sale_id, product_id, customer_id, reason, refund_amount, return_date))
                conn.execute("UPDATE products SET stock = stock + 1 WHERE id=?", (product_id,))
                conn.execute("UPDATE customers SET points = points - 10 WHERE id=?", (customer_id,))
                conn.execute("UPDATE warranties SET status='returned' WHERE sale_id=?", (sale_id,))
                conn.execute("""
                    UPDATE services SET status='cancelled'
                    WHERE product_id=? AND customer_id=? AND status='pending'
                """, (product_id, customer_id))
            else:
                flash("Sale not found.", "danger")
                return redirect("/returns")
        log_activity("RETURN", f"Return processed for sale {sale_id}")
        flash("Return processed and completed.", "success")
        return redirect("/returns")

    with get_db() as conn:
        returns = conn.execute("""
            SELECT r.id, p.name, c.name, r.reason, r.refund_amount, r.return_date, r.status
            FROM returns r
            JOIN products p ON r.product_id = p.id
            JOIN customers c ON r.customer_id = c.id
            ORDER BY r.return_date DESC
        """).fetchall()
        sales = conn.execute("SELECT id, product_id, total FROM sales").fetchall()
    return render_template("returns.html", returns=returns, sales=sales)

@app.route("/reports")
@login_required
def reports():
    with get_db() as conn:
        bar = conn.execute("""
            SELECT p.name, SUM(s.total) as total
            FROM sales s
            JOIN products p ON s.product_id = p.id
            GROUP BY p.name
        """).fetchall()
        labels = [row[0] for row in bar]
        totals = [row[1] for row in bar]
        pie = conn.execute("SELECT name, stock FROM products WHERE stock > 0").fetchall()
    return render_template("reports.html", labels=labels, totals=totals, pie_data=pie)

# ------------------ REORDER & QUICK RESTOCK ------------------
@app.route("/reorder")
@login_required
@admin_required
def reorder():
    with get_db() as conn:
        predictions = conn.execute("""
            SELECT p.id, p.name, p.stock,
                   COALESCE(ROUND(AVG(s.quantity) / 30.0, 2), 0) AS daily_avg
            FROM products p
            LEFT JOIN sales s ON p.id = s.product_id AND s.date >= date('now', '-30 days')
            GROUP BY p.id
        """).fetchall()
        reorder_items = []
        for row in predictions:
            pid, name, stock, daily_avg = row
            if daily_avg > 0 and stock < (daily_avg * 7):
                reorder_items.append({
                    'id': pid,
                    'name': name,
                    'stock': stock,
                    'daily_avg': daily_avg,
                    'days_remaining': int(stock / daily_avg) if daily_avg > 0 else 0
                })
    return render_template("reorder.html", items=reorder_items)

@app.route("/quick-restock/<int:product_id>")
@login_required
@admin_required
def quick_restock(product_id):
    """Add 10 units to a product's stock from the reorder page."""
    qty = 10
    notes = "Auto restock from reorder alert"
    date = datetime.now().strftime("%Y-%m-%d")
    with get_db() as conn:
        conn.execute("INSERT INTO stock_in (product_id, quantity, date, notes) VALUES (?, ?, ?, ?)",
                     (product_id, qty, date, notes))
        conn.execute("UPDATE products SET stock = stock + ? WHERE id = ?", (qty, product_id))
    log_activity("QUICK_RESTOCK", f"Added {qty} units to product ID {product_id} from reorder alert")
    flash(f"Added {qty} units to stock for product ID {product_id}.", "success")
    return redirect(url_for('reorder'))

# ------------------ TOOLS ------------------
@app.route("/tools")
@login_required
def tools():
    return render_template("tools.html")

@app.route("/backup-db")
@login_required
@admin_required
def backup_db():
    db_name = session.get("db_name")
    if not db_name:
        flash("No database selected.", "danger")
        return redirect("/tools")
    db_path = f"instance/{db_name}"
    if os.path.exists(db_path):
        return send_file(db_path, as_attachment=True, download_name=f"backup_{db_name}")
    else:
        flash("Database file not found.", "danger")
        return redirect("/tools")

@app.route("/export/<string:type>")
@login_required
def export(type):
    with get_db() as conn:
        wb = openpyxl.Workbook()
        ws = wb.active
        if type == "products":
            ws.title = "Products"
            data = conn.execute("""
                SELECT p.name, p.sku, p.price, p.stock, c.name, s.name
                FROM products p
                LEFT JOIN categories c ON p.category_id = c.id
                LEFT JOIN suppliers s ON p.supplier_id = s.id
            """).fetchall()
            headers = ["Name", "SKU", "Price", "Stock", "Category", "Supplier"]
        elif type == "customers":
            ws.title = "Customers"
            data = conn.execute("SELECT name, phone, email, address, points FROM customers").fetchall()
            headers = ["Name", "Phone", "Email", "Address", "Points"]
        elif type == "sales":
            ws.title = "Sales"
            data = conn.execute("""
                SELECT s.date, p.name, c.name, s.quantity, s.total, s.payment_method
                FROM sales s
                JOIN products p ON s.product_id = p.id
                LEFT JOIN customers c ON s.customer_id = c.id
                ORDER BY s.date DESC
            """).fetchall()
            headers = ["Date", "Product", "Customer", "Quantity", "Total", "Payment"]
        else:
            flash("Invalid export type", "danger")
            return redirect("/reports")

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row in data:
            ws.append(row)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(excel_file, download_name=f"{type}.xlsx", as_attachment=True)

@app.route("/import-products", methods=["POST"])
@login_required
@admin_required
def import_products():
    if "csv_file" not in request.files:
        flash("No file uploaded.", "danger")
        return redirect("/tools")
    file = request.files["csv_file"]
    if file.filename == "":
        flash("No file selected.", "danger")
        return redirect("/tools")
    if not file.filename.endswith(".csv"):
        flash("Please upload a CSV file.", "danger")
        return redirect("/tools")

    stream = StringIO(file.stream.read().decode("utf-8"))
    csv_reader = csv.reader(stream)
    next(csv_reader)
    count = 0
    with get_db() as conn:
        for row in csv_reader:
            if len(row) < 4:
                continue
            name = row[0].strip()
            price = float(row[1])
            category_name = row[2].strip()
            supplier_name = row[3].strip() if len(row) > 3 else ""

            cat = conn.execute("SELECT id FROM categories WHERE name=?", (category_name,)).fetchone()
            if not cat:
                conn.execute("INSERT INTO categories (name) VALUES (?)", (category_name,))
                cat = conn.execute("SELECT id FROM categories WHERE name=?", (category_name,)).fetchone()
            category_id = cat[0]

            supplier_id = None
            if supplier_name:
                sup = conn.execute("SELECT id FROM suppliers WHERE name=?", (supplier_name,)).fetchone()
                if not sup:
                    conn.execute("INSERT INTO suppliers (name, phone) VALUES (?, '')", (supplier_name,))
                    sup = conn.execute("SELECT id FROM suppliers WHERE name=?", (supplier_name,)).fetchone()
                supplier_id = sup[0]

            sku = name[:3].upper() + str(int(datetime.now().timestamp()))[-4:] + str(count)
            conn.execute("""
                INSERT INTO products (name, sku, category_id, supplier_id, price, stock)
                VALUES (?, ?, ?, ?, ?, 0)
            """, (name, sku, category_id, supplier_id, price))
            count += 1
    log_activity("IMPORT_PRODUCTS", f"Imported {count} products")
    flash(f"Imported {count} products successfully!", "success")
    return redirect("/tools")

# ------------------ INVOICE ------------------
@app.route("/invoice/<int:sale_id>")
@login_required
def invoice(sale_id):
    with get_db() as conn:
        sale = conn.execute("""
            SELECT s.total, s.quantity, s.warranty_months, s.date, s.payment_method,
                   p.name AS product_name,
                   c.name AS customer_name, c.phone, c.address, c.email
            FROM sales s
            JOIN products p ON s.product_id = p.id
            LEFT JOIN customers c ON s.customer_id = c.id
            WHERE s.id = ?
        """, (sale_id,)).fetchone()
        if not sale:
            flash("Sale not found", "danger")
            return redirect("/sales")

        total, qty, warranty_months, sale_date, payment_method, product_name, customer_name, phone, address, email = sale
        if not customer_name:
            customer_name = "Walk‑in Customer"

        qr_url = f"{request.url_root}warranty?search=sale_{sale_id}"
        qr_img = qrcode.make(qr_url)
        buffered = BytesIO()
        qr_img.save(buffered, format="PNG")
        qr_base64 = base64.b64encode(buffered.getvalue()).decode()

        return render_template("invoice.html",
                               invoice_id=sale_id,
                               date=sale_date,
                               customer_name=customer_name,
                               phone=phone,
                               email=email,
                               address=address,
                               product_name=product_name,
                               quantity=qty,
                               total=total,
                               warranty=warranty_months,
                               payment_method=payment_method or "cash",
                               qr_code=qr_base64)

# ------------------ LOGOUT ------------------
@app.route("/logout")
def logout():
    if "user" in session:
        log_activity("LOGOUT", "User logged out")
    session.clear()
    flash("You have been logged out.", "info")
    return redirect("/staff-login")

if __name__ == "__main__":
    os.makedirs("instance", exist_ok=True)
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)