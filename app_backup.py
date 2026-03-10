import os
import sqlite3
import bcrypt
import csv
import shutil
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from functools import wraps
import glob
from flask import Flask, render_template, request, redirect, session, flash, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO, StringIO

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "supersecretkey123")
app.permanent_session_lifetime = 3600  # 1 hour

# ------------------ HELPER FUNCTIONS ------------------
def get_db():
    """Return connection to the active company database."""
    db_name = session.get("db_name")
    if not db_name:
        return sqlite3.connect("instance/default.db")
    return sqlite3.connect(f"instance/{db_name}")

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            flash("Please log in first.", "warning")
            return redirect("/staff-login")
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Access denied. Admin only.", "danger")
            return redirect("/dashboard")
        return f(*args, **kwargs)
    return decorated_function

def log_activity(action, details=""):
    """Log user activity – uses its own connection, so call it outside other db blocks."""
    if "user" not in session:
        return
    with get_db() as conn:
        conn.execute("""
            INSERT INTO activity_log (user_id, username, action, details, timestamp)
            VALUES (?, ?, ?, ?, ?)
        """, (
            session.get("user_id"),
            session.get("user"),
            action,
            details,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
def send_email(recipient, subject, body):
    """Send email using company SMTP settings and environment password."""
    with get_db() as conn:
        settings = conn.execute("SELECT email_host, email_port, email_user, email_from FROM company").fetchone()
    if not settings or not settings[0]:
        print("❌ Email error: SMTP settings not found in database")
        return False
    host, port, user, from_addr = settings
    password = os.environ.get("EMAIL_PASSWORD")
    if not password:
        print("❌ Email error: EMAIL_PASSWORD environment variable not set")
        return False
    print(f"📧 Attempting to send email via {host}:{port} as {user}")
    msg = MIMEMultipart()
    msg['From'] = from_addr or user
    msg['To'] = recipient
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))
    try:
        server = smtplib.SMTP(host, port)
        server.starttls()
        server.login(user, password)
        server.send_message(msg)
        server.quit()
        print(f"✅ Email sent successfully to {recipient}")
        return True
    except Exception as e:
        print(f"❌ Email error: {e}")
        return False
def get_company_list():
    """Return list of existing company databases."""
    db_files = glob.glob("instance/stockify_*.db")
    companies = []
    for f in db_files:
        name = f[18:-3].replace('_', ' ').title()
        companies.append((os.path.basename(f), name))
    return companies

def init_company_db(db_name):
    """Create all tables for a new company database (with email columns)."""
    with sqlite3.connect(f"instance/{db_name}") as conn:
        # Users
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                password BLOB,
                role TEXT DEFAULT 'staff',
                email TEXT
            )
        """)
        try:
            conn.execute("ALTER TABLE users ADD COLUMN email TEXT")
        except sqlite3.OperationalError:
            pass

        # Company info (with email settings)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS company (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                gst TEXT,
                category TEXT,
                email_host TEXT,
                email_port INTEGER,
                email_user TEXT,
                email_password TEXT,
                email_from TEXT
            )
        """)
        # Add email columns if they don't exist (for older DBs)
        try:
            conn.execute("ALTER TABLE company ADD COLUMN email_host TEXT")
            conn.execute("ALTER TABLE company ADD COLUMN email_port INTEGER")
            conn.execute("ALTER TABLE company ADD COLUMN email_user TEXT")
            conn.execute("ALTER TABLE company ADD COLUMN email_password TEXT")
            conn.execute("ALTER TABLE company ADD COLUMN email_from TEXT")
        except sqlite3.OperationalError:
            pass

        # Categories
        conn.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE
            )
        """)

        # Suppliers
        conn.execute("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                phone TEXT
            )
        """)

        # Products (no barcode)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                sku TEXT UNIQUE,
                category_id INTEGER,
                supplier_id INTEGER,
                price REAL,
                stock INTEGER DEFAULT 0
            )
        """)

        # Stock movements
        conn.execute("""
            CREATE TABLE IF NOT EXISTS stock_in (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                quantity INTEGER,
                date TEXT,
                notes TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS stock_out (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                quantity INTEGER,
                date TEXT,
                notes TEXT
            )
        """)

        # Sales
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                customer_id INTEGER,
                quantity INTEGER,
                total REAL,
                date TEXT,
                warranty_months INTEGER DEFAULT 0
            )
        """)

        # Customers
        conn.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                phone TEXT UNIQUE,
                email TEXT,
                address TEXT
            )
        """)

        # Warranties
        conn.execute("""
            CREATE TABLE IF NOT EXISTS warranties (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER,
                product_id INTEGER,
                customer_id INTEGER,
                start_date TEXT,
                end_date TEXT,
                status TEXT DEFAULT 'active'
            )
        """)

        # Services
        conn.execute("""
            CREATE TABLE IF NOT EXISTS services (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                customer_id INTEGER,
                issue TEXT,
                service_date TEXT,
                resolution TEXT,
                cost REAL,
                status TEXT DEFAULT 'pending'
            )
        """)

        # Returns
        conn.execute("""
            CREATE TABLE IF NOT EXISTS returns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER,
                product_id INTEGER,
                customer_id INTEGER,
                reason TEXT,
                refund_amount REAL,
                return_date TEXT,
                status TEXT DEFAULT 'pending'
            )
        """)

        # Activity Log
        conn.execute("""
            CREATE TABLE IF NOT EXISTS activity_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                action TEXT,
                details TEXT,
                timestamp TEXT
            )
        """)

        # Insert default admin
        admin = conn.execute("SELECT * FROM users WHERE username='admin'").fetchone()
        if not admin:
            hashed = bcrypt.hashpw("admin".encode(), bcrypt.gensalt())
            conn.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                         ("admin", hashed, "admin"))

        # Default categories
        default_cats = ["Electronics", "Clothing", "Vehicles", "Hardware", "Food Items"]
        for cat in default_cats:
            conn.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (cat,))

# ------------------ ROUTES ------------------
@app.route("/")
def splash():
    return render_template("splash.html")

@app.route("/admin-login")
def admin_login():
    return render_template("admin_login.html")

@app.route("/staff-login")
def staff_login():
    companies = get_company_list()
    return render_template("staff_login.html", companies=companies)

@app.route("/login", methods=["POST"])
def login():
    if request.method == "POST":
        if "company_name" in request.form:
            # Admin login
            company = request.form["company_name"].strip()
            gst = request.form["gst"].strip()
            category = request.form["category"].strip()
            username = request.form["username"].strip()
            password = request.form["password"].strip()
            db_name = f"stockify_{company.replace(' ', '_').lower()}.db"
            session["db_name"] = db_name
            init_company_db(db_name)

            with sqlite3.connect(f"instance/{db_name}") as conn:
                conn.execute("DELETE FROM company")
                conn.execute("INSERT INTO company (name, gst, category) VALUES (?, ?, ?)",
                             (company, gst, category))
                user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()

            if user and bcrypt.checkpw(password.encode(), user[2]):
                session.permanent = True
                session["user"] = username
                session["user_id"] = user[0]
                session["company"] = company
                session["role"] = user[3]
                log_activity("LOGIN", "Admin login successful")
                flash(f"Welcome admin {username}!", "success")
                return redirect("/dashboard")
            else:
                flash("Invalid admin credentials", "danger")
                return redirect("/admin-login")
        else:
            # Staff login
            db_name = request.form["company_db"]
            username = request.form["username"].strip()
            password = request.form["password"].strip()
            session["db_name"] = db_name

            with sqlite3.connect(f"instance/{db_name}") as conn:
                user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
                if user and bcrypt.checkpw(password.encode(), user[2]):
                    company = conn.execute("SELECT name FROM company").fetchone()
                    session.permanent = True
                    session["user"] = username
                    session["user_id"] = user[0]
                    session["company"] = company[0] if company else "Unknown"
                    session["role"] = user[3]
                    log_activity("LOGIN", "Staff login successful")
                    flash(f"Welcome {username}!", "success")
                    return redirect("/dashboard")
                else:
                    flash("Invalid staff credentials", "danger")
                    return redirect("/staff-login")
    return redirect("/")

@app.route("/dashboard")
@login_required
def dashboard():
    with get_db() as conn:
        total_products = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        total_sales = conn.execute("SELECT SUM(total) FROM sales").fetchone()[0] or 0
        low_stock = conn.execute("SELECT COUNT(*) FROM products WHERE stock < 5").fetchone()[0]
        pending_services = conn.execute("SELECT COUNT(*) FROM services WHERE status='pending'").fetchone()[0]
        active_warranties = conn.execute("SELECT COUNT(*) FROM warranties WHERE status='active'").fetchone()[0]
        today = datetime.now().strftime("%Y-%m-%d")
        today_sales = conn.execute("SELECT SUM(total) FROM sales WHERE date=?", (today,)).fetchone()[0] or 0
        expiring = conn.execute("""
            SELECT COUNT(*) FROM warranties
            WHERE status='active' AND date(end_date) BETWEEN date('now') AND date('now', '+7 days')
        """).fetchone()[0]
    return render_template("dashboard.html",
                           total_products=total_products,
                           total_sales=total_sales,
                           low_stock_count=low_stock,
                           pending_services=pending_services,
                           active_warranties=active_warranties,
                           today_sales=today_sales,
                           expiring_warranties=expiring)

# ------------------ ADMIN ONLY ROUTES ------------------
@app.route("/products", methods=["GET", "POST"])
@login_required
@admin_required
def products():
    if request.method == "POST":
        # Add new product
        name = request.form["name"]
        price = float(request.form["price"])
        category = request.form["category"]
        supplier = request.form["supplier"]
        sku = name[:3].upper() + str(int(datetime.now().timestamp()))[-4:]
        with get_db() as conn:
            conn.execute("""
                INSERT INTO products (name, sku, category_id, supplier_id, price, stock)
                VALUES (?, ?, ?, ?, ?, 0)
            """, (name, sku, category, supplier, price))
        log_activity("ADD_PRODUCT", f"Added product: {name}")
        flash("Product added successfully!", "success")

    with get_db() as conn:
        categories = conn.execute("SELECT * FROM categories").fetchall()
        suppliers = conn.execute("SELECT * FROM suppliers").fetchall()
        products = conn.execute("""
            SELECT p.id, p.name, p.sku, p.price, p.stock,
                   c.name AS cat_name, s.name AS sup_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN suppliers s ON p.supplier_id = s.id
        """).fetchall()
    return render_template("products.html", products=products,
                           categories=categories, suppliers=suppliers)

@app.route("/add-stock/<int:product_id>", methods=["POST"])
@login_required
@admin_required
def add_stock(product_id):
    qty = int(request.form["quantity"])
    notes = request.form.get("notes", "")
    date = datetime.now().strftime("%Y-%m-%d")
    with get_db() as conn:
        conn.execute("INSERT INTO stock_in (product_id, quantity, date, notes) VALUES (?, ?, ?, ?)",
                     (product_id, qty, date, notes))
        conn.execute("UPDATE products SET stock = stock + ? WHERE id = ?", (qty, product_id))
    log_activity("STOCK_IN", f"Added {qty} units to product ID {product_id}")
    flash(f"Added {qty} units to stock.", "success")
    return redirect("/products")

@app.route("/remove-stock/<int:product_id>", methods=["POST"])
@login_required
@admin_required
def remove_stock(product_id):
    qty = int(request.form["quantity"])
    notes = request.form.get("notes", "")
    date = datetime.now().strftime("%Y-%m-%d")
    with get_db() as conn:
        current = conn.execute("SELECT stock FROM products WHERE id=?", (product_id,)).fetchone()
        if current and current[0] >= qty:
            conn.execute("INSERT INTO stock_out (product_id, quantity, date, notes) VALUES (?, ?, ?, ?)",
                         (product_id, qty, date, notes))
            conn.execute("UPDATE products SET stock = stock - ? WHERE id = ?", (qty, product_id))
            log_activity("STOCK_OUT", f"Removed {qty} units from product ID {product_id}")
            flash(f"Removed {qty} units from stock.", "success")
        else:
            flash("Insufficient stock!", "danger")
    return redirect("/products")

@app.route("/suppliers", methods=["GET", "POST"])
@login_required
@admin_required
def suppliers():
    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        with get_db() as conn:
            conn.execute("INSERT INTO suppliers (name, phone) VALUES (?, ?)", (name, phone))
        log_activity("ADD_SUPPLIER", f"Added supplier: {name}")
        flash("Supplier added successfully!", "success")

    with get_db() as conn:
        suppliers = conn.execute("SELECT * FROM suppliers").fetchall()
    return render_template("suppliers.html", suppliers=suppliers)

@app.route("/admin/users", methods=["GET", "POST"])
@login_required
@admin_required
def admin_users():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        role = request.form["role"]
        email = request.form.get("email", "")
        hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
        try:
            with get_db() as conn:
                conn.execute("INSERT INTO users (username, password, role, email) VALUES (?, ?, ?, ?)",
                             (username, hashed, role, email))
            log_activity("ADD_USER", f"Created user: {username} with role {role}")
            flash(f"User {username} created.", "success")
        except sqlite3.IntegrityError:
            flash("Username already exists.", "danger")

    with get_db() as conn:
        users = conn.execute("SELECT id, username, role, email FROM users").fetchall()
    return render_template("admin_users.html", users=users)

@app.route("/email-settings", methods=["GET", "POST"])
@login_required
@admin_required
def email_settings():
    with get_db() as conn:
        if request.method == "POST":
            host = request.form["host"]
            port = int(request.form["port"])
            user = request.form["user"]
            from_addr = request.form["from"]
            print(f"Saved settings: host={host}, port={port}, user={user}, from={from_addr}")
            conn.execute("""
                UPDATE company SET email_host=?, email_port=?, email_user=?, email_from=?
            """, (host, port, user, from_addr))
            flash("Email settings saved. Password is taken from environment variable.", "success")
        settings = conn.execute("SELECT email_host, email_port, email_user, email_from FROM company").fetchone()
    return render_template("email_settings.html", settings=settings)
@app.route("/activity-log")
@login_required
@admin_required
def activity_log():
    with get_db() as conn:
        logs = conn.execute("""
            SELECT username, action, details, timestamp
            FROM activity_log
            ORDER BY timestamp DESC LIMIT 100
        """).fetchall()
    return render_template("activity_log.html", logs=logs)

# ------------------ SHARED ROUTES (Staff & Admin) ------------------
@app.route("/sales", methods=["GET", "POST"])
@login_required
def sales():
    if request.method == "POST":
        product_id = request.form["product"]
        qty = int(request.form["qty"])
        customer_id = request.form.get("customer")
        warranty_months = int(request.form.get("warranty_months", 0))

        with get_db() as conn:
            product = conn.execute("SELECT price, stock FROM products WHERE id=?", (product_id,)).fetchone()
            if not product:
                flash("Product not found.", "danger")
                return redirect("/sales")
            price, stock = product
            if qty > stock:
                qty = stock
                flash(f"Quantity reduced to available stock ({stock}).", "warning")
            total = qty * price
            date = datetime.now().strftime("%Y-%m-%d")
            cur = conn.execute("""
                INSERT INTO sales (product_id, customer_id, quantity, total, date, warranty_months)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (product_id, customer_id, qty, total, date, warranty_months))
            sale_id = cur.lastrowid
            conn.execute("UPDATE products SET stock = stock - ? WHERE id=?", (qty, product_id))

            if warranty_months > 0 and customer_id:
                start_date = date
                end_date = (datetime.now() + timedelta(days=30 * warranty_months)).strftime("%Y-%m-%d")
                conn.execute("""
                    INSERT INTO warranties (sale_id, product_id, customer_id, start_date, end_date, status)
                    VALUES (?, ?, ?, ?, ?, 'active')
                """, (sale_id, product_id, customer_id, start_date, end_date))

            # Send email notification if customer has email
            if customer_id:
                customer = conn.execute("SELECT email, name FROM customers WHERE id=?", (customer_id,)).fetchone()
                if customer and customer[0]:
                    product_name = conn.execute("SELECT name FROM products WHERE id=?", (product_id,)).fetchone()[0]
                    subject = f"Your Order Confirmation – {product_name}"
                    body = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif;">
                        <h2>Thank you for your purchase, {customer[1]}!</h2>
                        <p>Here are your order details:</p>
                        <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">
                            <tr><th>Product</th><td>{product_name}</td></tr>
                            <tr><th>Quantity</th><td>{qty}</td></tr>
                            <tr><th>Total Amount</th><td>₹{total}</td></tr>
                            <tr><th>Warranty</th><td>{warranty_months} months</td></tr>
                            <tr><th>Order Date</th><td>{date}</td></tr>
                        </table>
                        <p>We hope you enjoy your purchase!</p>
                        <p>– Stockify Team</p>
                    </body>
                    </html>
                    """
                    sent = send_email(customer[0], subject, body)
                    if sent:
                        flash("Email confirmation sent to customer.", "success")
                    else:
                        flash("Email not sent – SMTP not configured or invalid.", "warning")

        log_activity("SALE", f"Sale of {qty} units of product ID {product_id} to customer {customer_id}")
        flash("Sale recorded successfully!", "success")

    with get_db() as conn:
        products = conn.execute("SELECT id, name, price, stock FROM products").fetchall()
        customers = conn.execute("SELECT id, name, phone FROM customers").fetchall()
        recent_sales = conn.execute("""
            SELECT s.date, p.name, c.name, s.quantity, s.total
            FROM sales s
            JOIN products p ON s.product_id = p.id
            LEFT JOIN customers c ON s.customer_id = c.id
            ORDER BY s.date DESC LIMIT 10
        """).fetchall()
    return render_template("sales.html", products=products, customers=customers, recent_sales=recent_sales)

@app.route("/customers", methods=["GET", "POST"])
@login_required
def customers():
    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        email = request.form.get("email", "")
        address = request.form.get("address", "")
        try:
            with get_db() as conn:
                conn.execute("INSERT INTO customers (name, phone, email, address) VALUES (?, ?, ?, ?)",
                             (name, phone, email, address))
            log_activity("ADD_CUSTOMER", f"Added customer: {name}")
            flash("Customer added successfully!", "success")
        except sqlite3.IntegrityError:
            flash("Phone number already exists.", "danger")

    with get_db() as conn:
        customers = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    return render_template("customers.html", customers=customers)

@app.route("/warranty")
@login_required
def warranty():
    with get_db() as conn:
        warranties = conn.execute("""
            SELECT w.id, p.name AS product, c.name AS customer,
                   w.start_date, w.end_date, w.status
            FROM warranties w
            JOIN products p ON w.product_id = p.id
            JOIN customers c ON w.customer_id = c.id
            ORDER BY w.end_date
        """).fetchall()
    return render_template("warranty.html", warranties=warranties)

@app.route("/service", methods=["GET", "POST"])
@login_required
def service():
    if request.method == "POST":
        product_id = request.form["product"]
        customer_id = request.form["customer"]
        issue = request.form["issue"]
        service_date = datetime.now().strftime("%Y-%m-%d")
        with get_db() as conn:
            conn.execute("""
                INSERT INTO services (product_id, customer_id, issue, service_date, status)
                VALUES (?, ?, ?, ?, 'pending')
            """, (product_id, customer_id, issue, service_date))
        log_activity("SERVICE_CREATED", f"Service logged for product ID {product_id}")
        flash("Service request logged.", "success")

    with get_db() as conn:
        products = conn.execute("SELECT id, name FROM products").fetchall()
        customers = conn.execute("SELECT id, name FROM customers").fetchall()
        services = conn.execute("""
            SELECT s.id, p.name, c.name, s.issue, s.service_date, s.status
            FROM services s
            JOIN products p ON s.product_id = p.id
            JOIN customers c ON s.customer_id = c.id
            ORDER BY s.service_date DESC
        """).fetchall()
    return render_template("service.html", products=products, customers=customers, services=services)

@app.route("/service/update/<int:service_id>", methods=["POST"])
@login_required
def update_service(service_id):
    status = request.form["status"]
    resolution = request.form.get("resolution", "")
    cost = float(request.form.get("cost", 0))
    with get_db() as conn:
        conn.execute("""
            UPDATE services SET status=?, resolution=?, cost=?
            WHERE id=?
        """, (status, resolution, cost, service_id))
    log_activity("SERVICE_UPDATED", f"Service {service_id} updated to {status}")
    flash("Service updated.", "success")
    return redirect("/service")

@app.route("/returns", methods=["GET", "POST"])
@login_required
def returns():
    if request.method == "POST":
        sale_id = request.form["sale_id"]
        reason = request.form["reason"]
        refund_amount = float(request.form["refund_amount"])
        return_date = datetime.now().strftime("%Y-%m-%d")

        with get_db() as conn:
            sale = conn.execute("SELECT product_id, customer_id FROM sales WHERE id=?", (sale_id,)).fetchone()
            if sale:
                product_id, customer_id = sale
                conn.execute("""
                    INSERT INTO returns (sale_id, product_id, customer_id, reason, refund_amount, return_date, status)
                    VALUES (?, ?, ?, ?, ?, ?, 'pending')
                """, (sale_id, product_id, customer_id, reason, refund_amount, return_date))
                # Restock 1 unit (assuming one unit returned)
                conn.execute("UPDATE products SET stock = stock + 1 WHERE id=?", (product_id,))
            else:
                flash("Sale not found.", "danger")
                return redirect("/returns")
        log_activity("RETURN", f"Return processed for sale {sale_id}")
        flash("Return processed.", "success")

    with get_db() as conn:
        returns = conn.execute("""
            SELECT r.id, p.name, c.name, r.reason, r.refund_amount, r.return_date, r.status
            FROM returns r
            JOIN products p ON r.product_id = p.id
            JOIN customers c ON r.customer_id = c.id
            ORDER BY r.return_date DESC
        """).fetchall()
        sales = conn.execute("SELECT id, product_id, total FROM sales").fetchall()
    return render_template("returns.html", returns=returns, sales=sales)

@app.route("/reports")
@login_required
def reports():
    with get_db() as conn:
        bar = conn.execute("""
            SELECT p.name, SUM(s.total) as total
            FROM sales s
            JOIN products p ON s.product_id = p.id
            GROUP BY p.name
        """).fetchall()
        labels = [row[0] for row in bar]
        totals = [row[1] for row in bar]
        pie = conn.execute("SELECT name, stock FROM products WHERE stock > 0").fetchall()
    return render_template("reports.html", labels=labels, totals=totals, pie_data=pie)

# ------------------ TOOLS ROUTES ------------------
@app.route("/tools")
@login_required
def tools():
    return render_template("tools.html")

@app.route("/backup-db")
@login_required
@admin_required
def backup_db():
    db_name = session.get("db_name")
    if not db_name:
        flash("No database selected.", "danger")
        return redirect("/tools")
    db_path = f"instance/{db_name}"
    if os.path.exists(db_path):
        return send_file(db_path, as_attachment=True, download_name=f"backup_{db_name}")
    else:
        flash("Database file not found.", "danger")
        return redirect("/tools")

@app.route("/export/<string:type>")
@login_required
def export(type):
    with get_db() as conn:
        wb = openpyxl.Workbook()
        ws = wb.active
        if type == "products":
            ws.title = "Products"
            data = conn.execute("""
                SELECT p.name, p.sku, p.price, p.stock, c.name, s.name
                FROM products p
                LEFT JOIN categories c ON p.category_id = c.id
                LEFT JOIN suppliers s ON p.supplier_id = s.id
            """).fetchall()
            headers = ["Name", "SKU", "Price", "Stock", "Category", "Supplier"]
        elif type == "customers":
            ws.title = "Customers"
            data = conn.execute("SELECT name, phone, email, address FROM customers").fetchall()
            headers = ["Name", "Phone", "Email", "Address"]
        elif type == "sales":
            ws.title = "Sales"
            data = conn.execute("""
                SELECT s.date, p.name, c.name, s.quantity, s.total
                FROM sales s
                JOIN products p ON s.product_id = p.id
                LEFT JOIN customers c ON s.customer_id = c.id
                ORDER BY s.date DESC
            """).fetchall()
            headers = ["Date", "Product", "Customer", "Quantity", "Total"]
        else:
            flash("Invalid export type", "danger")
            return redirect("/reports")

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row in data:
            ws.append(row)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(excel_file, download_name=f"{type}.xlsx", as_attachment=True)

@app.route("/import-products", methods=["POST"])
@login_required
@admin_required
def import_products():
    if "csv_file" not in request.files:
        flash("No file uploaded.", "danger")
        return redirect("/tools")
    file = request.files["csv_file"]
    if file.filename == "":
        flash("No file selected.", "danger")
        return redirect("/tools")
    if not file.filename.endswith(".csv"):
        flash("Please upload a CSV file.", "danger")
        return redirect("/tools")

    stream = StringIO(file.stream.read().decode("utf-8"))
    csv_reader = csv.reader(stream)
    next(csv_reader)  # skip header
    count = 0
    with get_db() as conn:
        for row in csv_reader:
            if len(row) < 4:
                continue
            name = row[0].strip()
            price = float(row[1])
            category_name = row[2].strip()
            supplier_name = row[3].strip() if len(row) > 3 else ""

            # Get or create category
            cat = conn.execute("SELECT id FROM categories WHERE name=?", (category_name,)).fetchone()
            if not cat:
                conn.execute("INSERT INTO categories (name) VALUES (?)", (category_name,))
                cat = conn.execute("SELECT id FROM categories WHERE name=?", (category_name,)).fetchone()
            category_id = cat[0]

            # Get or create supplier
            supplier_id = None
            if supplier_name:
                sup = conn.execute("SELECT id FROM suppliers WHERE name=?", (supplier_name,)).fetchone()
                if not sup:
                    conn.execute("INSERT INTO suppliers (name, phone) VALUES (?, '')", (supplier_name,))
                    sup = conn.execute("SELECT id FROM suppliers WHERE name=?", (supplier_name,)).fetchone()
                supplier_id = sup[0]

            sku = name[:3].upper() + str(int(datetime.now().timestamp()))[-4:] + str(count)
            conn.execute("""
                INSERT INTO products (name, sku, category_id, supplier_id, price, stock)
                VALUES (?, ?, ?, ?, ?, 0)
            """, (name, sku, category_id, supplier_id, price))
            count += 1
    log_activity("IMPORT_PRODUCTS", f"Imported {count} products")
    flash(f"Imported {count} products successfully!", "success")
    return redirect("/tools")

@app.route("/logout")
def logout():
    if "user" in session:
        log_activity("LOGOUT", "User logged out")
    session.clear()
    flash("You have been logged out.", "info")
    return redirect("/staff-login")

if __name__ == "__main__":
    os.makedirs("instance", exist_ok=True)
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)